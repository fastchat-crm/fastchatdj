"""
Importa `PLANES INDIVIDUALES VIDA SANA 2026.xlsx` a las tablas del cotizador.

Idempotente: se puede correr varias veces sin duplicar (usa update_or_create
sobre claves naturales).

Uso:
    python manage.py import_excel_vidabuena --archivo "/ruta/PLANES INDIVIDUALES VIDA SANA 2026.xlsx" --perfil-id 1
    python manage.py import_excel_vidabuena --archivo ... --usuario-email asesor@vidabuena.ec

El --perfil-id / --usuario-email indica a qué PerfilNegocioIA (empresa) se cargan
los planes (multi-tenant).
"""
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from crm.models import PerfilNegocioIA
from cotizador.models import (
    VigenciaTarifaria, RangoEtario, Plan, Tarifa, Cobertura, ProcedimientoDental,
)

# Hoja de tarifa  ->  (nombre_comercial, codigo, columna en CUADRO DE COBERTURA)
HOJAS_PLAN = {
    'PROTECCIÓN 10,000': ('Protección 10.000', 'PROTECCION_10000', 4),
    'UNICO 10,000':      ('Único 10.000',      'UNICO_10000',      5),
    'PREDILECTO 20,000': ('Predilecto 20.000', 'PREDILECTO_20000', 6),
    'MAGNO 30,000':      ('Magno 30.000',      'MAGNO_30000',      7),
}

# Columnas de las hojas de tarifa (1-based): A edad, B/C/D/E primas
COL_EDAD = 1
COL_BASICO_M, COL_PLUS_M, COL_BASICO_F, COL_PLUS_F = 2, 3, 4, 5

MODALIDAD_MAP = {'mixta': 'mixta', 'cerrada': 'cerrada'}
TIPO_COB_MAP = {'anual': 'anual', 'por incapacidad': 'por_incapacidad'}


def _dec(valor):
    if valor is None:
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        return None


def _parse_rango(etiqueta):
    """'(20 - 25)' -> (20, 25, '(20 - 25)')"""
    m = re.findall(r'\d+', str(etiqueta or ''))
    if len(m) >= 2:
        return int(m[0]), int(m[1]), str(etiqueta).strip()
    return None


def _parse_deducible(texto):
    if texto is None:
        return Decimal('0')
    t = str(texto).strip().lower()
    if 'sin' in t:
        return Decimal('0')
    nums = re.findall(r'\d+(?:[.,]\d+)?', t)
    return _dec(nums[0].replace(',', '.')) if nums else Decimal('0')


def _parse_dias(texto):
    nums = re.findall(r'\d+', str(texto or ''))
    return int(nums[0]) if nums else None


class Command(BaseCommand):
    help = 'Importa el Excel de planes y tarifas de Vida Buena a la BD del cotizador.'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', required=True, help='Ruta al .xlsx')
        parser.add_argument('--perfil-id', type=int, help='ID de PerfilNegocioIA (empresa)')
        parser.add_argument('--usuario-email', help='Email del usuario dueño del perfil')
        parser.add_argument('--vigencia', default='Tarifario 2026', help='Nombre de la vigencia')

    def _get_empresa(self, opts):
        if opts.get('perfil_id'):
            try:
                return PerfilNegocioIA.objects.get(id=opts['perfil_id'])
            except PerfilNegocioIA.DoesNotExist:
                raise CommandError(f"No existe PerfilNegocioIA id={opts['perfil_id']}")
        if opts.get('usuario_email'):
            try:
                return PerfilNegocioIA.objects.get(usuario__email=opts['usuario_email'])
            except PerfilNegocioIA.DoesNotExist:
                raise CommandError(f"No hay perfil para {opts['usuario_email']}")
        raise CommandError('Indica --perfil-id o --usuario-email')

    @transaction.atomic
    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Falta openpyxl: pip install openpyxl')

        empresa = self._get_empresa(opts)
        wb = openpyxl.load_workbook(opts['archivo'], data_only=True)

        vigencia, _ = VigenciaTarifaria.objects.update_or_create(
            empresa=empresa, nombre=opts['vigencia'],
            defaults={'activa': True},
        )

        # 1) Rangos etarios (desde la hoja PROTECCIÓN, col EDAD desde fila 5)
        rangos = {}
        ws0 = wb['PROTECCIÓN 10,000']
        for row in range(5, ws0.max_row + 1):
            etq = ws0.cell(row=row, column=COL_EDAD).value
            parsed = _parse_rango(etq)
            if not parsed:
                continue
            emin, emax, label = parsed
            rango, _ = RangoEtario.objects.update_or_create(
                empresa=empresa, etiqueta=label,
                defaults={'edad_min': emin, 'edad_max': emax},
            )
            rangos[label] = rango
        self.stdout.write(f'Rangos etarios: {len(rangos)}')

        # 2) Planes + condiciones generales (CUADRO DE COBERTURA filas 11-16)
        cuadro = wb['CUADRO DE COBERTURA']
        planes = {}
        for hoja, (nombre, codigo, col) in HOJAS_PLAN.items():
            nivel = cuadro.cell(row=11, column=col).value
            modalidad = str(cuadro.cell(row=12, column=col).value or '').strip().lower()
            tipo_cob = str(cuadro.cell(row=13, column=col).value or '').strip().lower()
            deducible = _parse_deducible(cuadro.cell(row=14, column=col).value)
            territorio = cuadro.cell(row=15, column=col).value
            periodo = _parse_dias(cuadro.cell(row=16, column=col).value)
            suma = _dec(re.sub(r'\D', '', codigo.split('_')[-1]))

            plan, _ = Plan.objects.update_or_create(
                empresa=empresa, codigo=codigo,
                defaults={
                    'nombre_comercial': nombre,
                    'suma_asegurada': suma,
                    'modalidad': MODALIDAD_MAP.get(modalidad),
                    'tipo_cobertura': TIPO_COB_MAP.get(tipo_cob),
                    'nivel_referencia': str(nivel).strip() if nivel else None,
                    'deducible_anual': deducible or Decimal('0'),
                    'cobertura_territorial': str(territorio).strip() if territorio else None,
                    'periodo_presentacion_dias': periodo,
                    'orden': list(HOJAS_PLAN).index(hoja),
                },
            )
            planes[col] = plan
        self.stdout.write(f'Planes: {len(planes)}')

        # 3) Coberturas (CUADRO DE COBERTURA filas 18-108, con categoría por bloque [CAT])
        coberturas_creadas = 0
        categoria_actual = 'GENERAL'
        for row in range(18, cuadro.max_row + 1):
            concepto = cuadro.cell(row=row, column=2).value
            if not concepto:
                continue
            concepto = str(concepto).strip()
            valores = {col: cuadro.cell(row=row, column=col).value for col in planes}
            tiene_valor = any(v not in (None, '') for v in valores.values())
            if concepto.isupper() and not tiene_valor:
                categoria_actual = concepto[:80]
                continue
            for col, plan in planes.items():
                val = valores[col]
                if val in (None, ''):
                    continue
                Cobertura.objects.update_or_create(
                    plan=plan, categoria=categoria_actual, concepto=concepto[:255],
                    defaults={'valor': str(val)[:255], 'orden': row},
                )
                coberturas_creadas += 1
        self.stdout.write(f'Coberturas: {coberturas_creadas}')

        # 4) Tarifas (4 hojas)
        tarifas_creadas = 0
        for hoja, (nombre, codigo, col) in HOJAS_PLAN.items():
            ws = wb[hoja]
            plan = Plan.objects.get(empresa=empresa, codigo=codigo)
            for row in range(5, ws.max_row + 1):
                parsed = _parse_rango(ws.cell(row=row, column=COL_EDAD).value)
                if not parsed:
                    continue
                rango = rangos.get(parsed[2])
                if rango is None:
                    rango, _ = RangoEtario.objects.update_or_create(
                        empresa=empresa, etiqueta=parsed[2],
                        defaults={'edad_min': parsed[0], 'edad_max': parsed[1]},
                    )
                    rangos[parsed[2]] = rango
                celdas = {
                    ('M', 'basico'): ws.cell(row=row, column=COL_BASICO_M).value,
                    ('M', 'plus'):   ws.cell(row=row, column=COL_PLUS_M).value,
                    ('F', 'basico'): ws.cell(row=row, column=COL_BASICO_F).value,
                    ('F', 'plus'):   ws.cell(row=row, column=COL_PLUS_F).value,
                }
                for (genero, variante), prima in celdas.items():
                    prima_dec = _dec(prima)
                    if prima_dec is None:
                        continue
                    Tarifa.objects.update_or_create(
                        plan=plan, vigencia=vigencia, rango_etario=rango,
                        genero=genero, variante_dental=variante,
                        defaults={'prima_mensual': round(prima_dec, 2)},
                    )
                    tarifas_creadas += 1
        self.stdout.write(f'Tarifas: {tarifas_creadas}')

        # 5) Procedimientos dentales (2 hojas)
        dentales = 0
        mapa_dental = {'COBERTURA DENTAL BASICA': ('basico', 2, 3, None),
                       'COBERTURA DENTAL PLUS':  ('plus', 3, 4, 2)}
        for hoja, (variante, col_proc, col_copago, col_serv) in mapa_dental.items():
            if hoja not in wb.sheetnames:
                continue
            ws = wb[hoja]
            for row in range(4, ws.max_row + 1):
                proc = ws.cell(row=row, column=col_proc).value
                if not proc or not str(proc).strip():
                    continue
                copago = ws.cell(row=row, column=col_copago).value
                copago_dec = _dec(copago) if str(copago).strip() not in ('-', '') else None
                serv = ws.cell(row=row, column=col_serv).value if col_serv else None
                ProcedimientoDental.objects.update_or_create(
                    empresa=empresa, variante=variante, procedimiento=str(proc).strip()[:255],
                    defaults={
                        'copago': copago_dec,
                        'servicio': str(serv).strip()[:120] if serv else None,
                    },
                )
                dentales += 1
        self.stdout.write(f'Procedimientos dentales: {dentales}')

        self.stdout.write(self.style.SUCCESS(
            f'OK. Empresa #{empresa.id} — planes={len(planes)} tarifas={tarifas_creadas} '
            f'coberturas={coberturas_creadas} dentales={dentales}'
        ))
