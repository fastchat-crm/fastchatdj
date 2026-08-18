"""Exportación a Excel del listado de conversaciones caducadas.

Consumido por el action `exportar_excel` de `whatsapp/view_conversaciones.py`
(vista `/whatsapp/conversaciones-caducadas/`). Reutiliza el mismo `Q` de filtros
que arma el listado para que el archivo salga siempre alineado con lo que el
asesor ve en pantalla.
"""

from datetime import timedelta

from django.db.models import OuterRef, Subquery, F
from django.utils import timezone

from .models import ConversacionWhatsApp, MensajeWhatsApp

HORAS_VENTANA_META = 24

ETIQUETA_POR_TIPO = {
    'imagen': '[Imagen]',
    'video': '[Video]',
    'audio': '[Audio]',
    'documento': '[Documento]',
    'ubicacion': '[Ubicación]',
    'contacto': '[Contacto]',
    'sticker': '[Sticker]',
    'plantilla': '[Plantilla]',
}


def _texto_mensaje(texto, tipo):
    """Texto legible del mensaje: los adjuntos no traen cuerpo, se etiquetan."""
    texto = (texto or '').strip()
    if texto:
        return texto[:1000]
    return ETIQUETA_POR_TIPO.get(tipo or '', '')


def _fecha(valor):
    """El proyecto corre con USE_TZ=False (settings.py), así que los datetimes
    llegan naive y `localtime()` reventaría. Solo se convierte cuando el valor
    trae tzinfo."""
    if not valor:
        return ''
    if timezone.is_aware(valor):
        valor = timezone.localtime(valor)
    return valor.strftime('%d/%m/%Y %H:%M')


def _walink(numero):
    """URL wa.me a partir del número del contacto (solo dígitos)."""
    digitos = ''.join(c for c in (numero or '') if c.isdigit())
    if not digitos:
        return ''
    return f'https://wa.me/{digitos}'


def queryset_caducadas_export(filtros):
    """Conversaciones caducadas (ventana Meta de 24h vencida) con el último
    mensaje entrante y el último saliente resueltos por Subquery — mismo criterio
    que el branch `load_conversations` del listado, sin N+1."""
    ahora = timezone.now()

    entrante = (
        MensajeWhatsApp.objects
        .filter(conversacion=OuterRef('pk'))
        .exclude(remitente=OuterRef('contacto__sesion__numero'))
        .order_by('-fecha')
    )
    saliente = (
        MensajeWhatsApp.objects
        .filter(conversacion=OuterRef('pk'), remitente=OuterRef('contacto__sesion__numero'))
        .order_by('-fecha')
    )

    qs = (
        ConversacionWhatsApp.objects
        .filter(filtros, conversacion_finalizada=False)
        .select_related('contacto', 'contacto__sesion', 'asignado_a', 'primer_agente')
        .annotate(
            fecha_ultimo_entrante=Subquery(entrante.values('fecha')[:1]),
            texto_ultimo_entrante=Subquery(entrante.values('mensaje')[:1]),
            tipo_ultimo_entrante=Subquery(entrante.values('tipo')[:1]),
            fecha_ultima_respuesta=Subquery(saliente.values('fecha')[:1]),
            texto_ultima_respuesta=Subquery(saliente.values('mensaje')[:1]),
            tipo_ultima_respuesta=Subquery(saliente.values('tipo')[:1]),
        )
        .filter(
            contacto__sesion__proveedor='meta',
            fecha_ultimo_entrante__lte=ahora - timedelta(hours=HORAS_VENTANA_META),
        )
        .order_by(F('fecha_ultimo_entrante').desc(nulls_last=True), '-id')
        .distinct()
    )
    return qs


def exportar_caducadas_excel(qs):
    """Arma el Workbook de openpyxl del listado de conversaciones caducadas."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Caducadas'
    encabezados = [
        'Cod', 'Contacto', 'Número', 'WhatsApp', 'Sesión', 'Asesor asignado',
        'Última respuesta enviada', 'Fecha última respuesta',
        'Último mensaje del cliente', 'Fecha último mensaje del cliente',
        'Ventana Meta venció',
    ]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'

    for conv in qs:
        contacto = conv.contacto
        sesion = getattr(contacto, 'sesion', None)
        asesor = conv.asignado_a or conv.primer_agente
        numero = getattr(contacto, 'contacto_numero', '') or ''
        vence = None
        if conv.fecha_ultimo_entrante:
            vence = conv.fecha_ultimo_entrante + timedelta(hours=HORAS_VENTANA_META)

        ws.append([
            conv.id,
            getattr(contacto, 'contacto_nombre', '') or '',
            numero,
            'Abrir chat' if _walink(numero) else '',
            getattr(sesion, 'nombre', '') or getattr(sesion, 'numero', '') or '',
            (asesor.get_full_name() or asesor.username) if asesor else 'Sin asesor',
            _texto_mensaje(conv.texto_ultima_respuesta, conv.tipo_ultima_respuesta),
            _fecha(conv.fecha_ultima_respuesta),
            _texto_mensaje(conv.texto_ultimo_entrante, conv.tipo_ultimo_entrante),
            _fecha(conv.fecha_ultimo_entrante),
            _fecha(vence),
        ])

        enlace = _walink(numero)
        if enlace:
            celda = ws.cell(row=ws.max_row, column=4)
            celda.hyperlink = enlace
            celda.style = 'Hyperlink'

    anchos = [8, 28, 16, 14, 22, 24, 50, 20, 50, 20, 20]
    for indice, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=indice).column_letter].width = ancho

    return wb
