import sys
from datetime import date
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.template.loader import get_template
from openpyxl.reader.excel import load_workbook

from area_geografica.models import Ciudad, Provincia
from core.custom_models import FormError
from core.funciones import addData, paginador, salva_auditoria, secure_module, redirectAfterPostGet, codnombre, log
from core.funciones_adicionales import salva_logs
from .forms import UserForm, GrupoUserForm, ManageProfileForm, ChangeUsernameForm, CargarUsuariosForm

from django.contrib import messages

from .models import Usuario, PerfilAdministrativo, PerfilPersona
from .view_persona import personasView


PERMISOS_ESPECIALES = [
    {
        'app': 'crm',
        'codename': 'puede_ver_citas_all',
        'nombre': 'Ver todas las citas de la agenda',
        'descripcion': 'Sin este permiso el usuario solo ve las citas de los recursos que tiene asignados. Los superusuarios siempre ven todo.',
    },
]


def _permisos_especiales_de(usuario):
    from django.contrib.auth.models import Permission
    resultado = []
    for p in PERMISOS_ESPECIALES:
        asignado = usuario.user_permissions.filter(
            content_type__app_label=p['app'], codename=p['codename']
        ).exists()
        resultado.append({
            'clave': f"{p['app']}.{p['codename']}",
            'nombre': p['nombre'],
            'descripcion': p['descripcion'],
            'asignado': asignado,
        })
    return resultado


@login_required
@secure_module

def usuarioView(request):
    data = {
        'titulo': 'Control de Usuarios Administrativos',
        'descripcion': 'Crear, Editar y Eliminar Administradores',
        'modulo': 'Autenticación',
        'ruta': request.path,
        'fecha': str(date.today()),
    }
    addData(request, data)
    model = Usuario
    Formulario = UserForm

    if request.method == 'POST':
        res_json = []
        action = request.POST['action']
        try:
            with transaction.atomic():
                if action == 'add':
                    form = Formulario(request.POST, request.FILES, request=request)
                    if form.is_valid():
                        if 'ciudad' in request.POST:
                            form.instance.ciudad_id = request.POST['ciudad']
                        username_ = codnombre(request.POST['first_name'], request.POST['last_name'])
                        form.instance.username = username_
                        form.instance.set_password(request.POST["documento"])
                        obj = form.save()
                        perfil_ = PerfilAdministrativo(usuario_id=form.instance.id)
                        perfil_.save()
                        log(f"Habilito usuario {obj.username} - {obj.get_full_name()}", request, "add")
                        messages.success(request, "Agregado correctamente.")
                        res_json.append({'error': False,
                                         "to": redirectAfterPostGet(request)
                                         })
                    else:
                        raise FormError(form)
                elif action == 'change':

                    user = model.objects.get(pk=int(request.POST['pk']))

                    form = Formulario(request.POST, request.FILES, instance=user, request=request)
                    if form.is_valid():
                        if 'ciudad' in request.POST:
                            form.instance.ciudad_id = request.POST['ciudad']
                        form.save()
                        form.instance.get_admin()
                        log(f"Edito usuario {form.instance.username} - {form.instance.get_full_name()}", request,
                            "change")

                        messages.success(request, "Modificado correctamente.")
                        res_json.append({'error': False,
                                         "to": redirectAfterPostGet(request)
                                         })
                    else:
                        raise FormError(form)
                elif action == 'crearperfilpersona':
                    user = model.objects.get(pk=int(request.POST['id']))
                    group = Group.objects.get(name='Cliente')
                    user.groups.add(group)
                    if not user.es_persona():
                        if PerfilPersona.objects.filter(usuario=user).exists():
                            perfil_ = PerfilPersona.objects.get(usuario=user)
                            perfil_.status = True
                            perfil_.save(request)
                            log(f"Activo perfil cliente usuario {user.username} - {user.get_full_name()}", request, "change")
                            messages.success(request, "Perfil persona habilitado.")
                        else:
                            perfil_ = PerfilPersona(usuario=user)
                            perfil_.save(request)
                            log(f"Creo perfil cliente usuario {user.username} - {user.get_full_name()}", request, "add")
                            messages.success(request, "Perfil persona habilitado.")
                        res_json.append({'error': False, "reload": True})
                    else:
                        raise NameError(f"Usuario ya cuenta con perfil persona")
                elif action == 'permisos_especiales':
                    from django.contrib.auth.models import Permission
                    filtro = model.objects.get(pk=int(request.POST['pk']))
                    seleccion = request.POST.getlist('permisos')
                    for p in PERMISOS_ESPECIALES:
                        perm = Permission.objects.filter(
                            content_type__app_label=p['app'], codename=p['codename']
                        ).first()
                        if not perm:
                            res_json.append({
                                'error': True,
                                'message': f"El permiso {p['codename']} no existe todavía en la base — corre las migraciones de crm.",
                            })
                            return JsonResponse(res_json, safe=False)
                        clave = f"{p['app']}.{p['codename']}"
                        if clave in seleccion:
                            filtro.user_permissions.add(perm)
                        else:
                            filtro.user_permissions.remove(perm)
                    log(f"Permisos especiales actualizados para {filtro.username} - {filtro.get_full_name()}", request, "change")
                    res_json.append({'error': False, 'message': 'Permisos actualizados.', 'reload': True})
                elif action == 'changegroup':
                    filtro = model.objects.get(pk=int(request.POST['pk']))
                    form = GrupoUserForm(request.POST, request.FILES, instance=filtro, request=request)
                    if form.is_valid() and filtro:
                        form.save()
                        log(f"Cambio grupo usuario {filtro.username} - {filtro.get_full_name()}", request, "change")
                        res_json.append({'error': False, "reload": True})
                    else:
                        res_json.append({'error': True,
                                         "form": [{k: v[0]} for k, v in form.errors.items()],
                                         "message": "Error en el formulario"
                                         })
                elif action == 'delete':
                    user = model.objects.get(pk=int(request.POST['id']))
                    user.is_active = False
                    user.status = False
                    user.is_superuser = False
                    user.is_staff = False
                    user.save(request)
                    administrativo = user.get_admin()
                    administrativo.status = False
                    administrativo.save(request)
                    cliente = user.get_client()
                    cliente.status = False
                    cliente.save(request)
                    log(f"Inhabilito usuario {user.username} - {user.get_full_name()}", request, "del")
                    messages.success(request, "Inhabilitado correctamente.")
                    res_json = {"error": False}
                elif action == 'deleteperfiladm':
                    user = model.objects.get(pk=int(request.POST['id']))
                    perfil_ = PerfilPersona.objects.get(usuario=user)
                    perfil_.status = False
                    perfil_.save(request)
                    log(f"Inhabilitado perfil cliente {user.get_full_name()}", request, "del")
                    messages.success(request, "Perfil administrativo inhabilitado.")
                    res_json={"error":False}
                elif action == 'activate':
                    user = model.objects.get(pk=int(request.POST['id']))
                    user.is_active = True
                    user.is_staff = True
                    user.status = True
                    user.save(request)
                    log(f"Habilito usuario {user.username} - {user.get_full_name()}", request, "add")
                    messages.success(request, "Habilitado correctamente.")
                    res_json.append({'error': False,'reload': True})
                elif action == 'change_password':
                    user = model.objects.get(pk=int(request.POST['pk']))
                    user.set_password(request.POST["password"])
                    user.save(request)
                    log(f"Modifico contraseña usuario {user.username} - {user.get_full_name()}", request, "change")

                    messages.success(request, "Contraseña del usuario {} / {} cambiada".format(user.get_full_name(),
                                                                                               user.username))
                    res_json.append({'error': False,'reload': True})
                elif action == 'change_password_masivo':
                    # Cambia la contraseña de TODOS los usuarios que matchean el
                    # filtro actual del listado (el modal reenvía el querystring
                    # en filtros_qs). El propio usuario se excluye para no
                    # invalidar su sesión; los superusuarios solo los puede
                    # tocar otro superusuario.
                    from django.http import QueryDict
                    from .funciones_usuario import filtros_listado_usuarios
                    password = (request.POST.get('password') or '').strip()
                    if len(password) < 6:
                        raise ValueError('La contraseña debe tener al menos 6 caracteres.')
                    params = QueryDict((request.POST.get('filtros_qs') or '').lstrip('?'))
                    filtros_masivo = filtros_listado_usuarios(params)
                    usuarios_masivo = (
                        model.objects.filter(filtros_masivo)
                        .filter(perfiladministrativo__isnull=False)
                        .exclude(pk=request.user.pk)
                        .distinct()
                    )
                    if not request.user.is_superuser:
                        usuarios_masivo = usuarios_masivo.exclude(is_superuser=True)
                    total_masivo = 0
                    for usuario_m in usuarios_masivo:
                        usuario_m.set_password(password)
                        usuario_m.save(request)
                        total_masivo += 1
                    log(f"Cambio masivo de contraseña a {total_masivo} usuarios según filtro ({params.urlencode()})",
                        request, "change")
                    messages.success(request,
                                     f"Contraseña actualizada para {total_masivo} usuarios. "
                                     "Tu propio usuario queda excluido.")
                    res_json.append({'error': False, 'reload': True})
                elif action == 'eliminar_foto':
                    user = model.objects.get(pk=int(request.POST['pk']))
                    user.foto = ""
                    user.save(request)
                    log(f"Elimino usuario {user.username} - {user.get_full_name()}", request, "del")
                    return JsonResponse({'state': True})
                elif action == 'manage_profile':
                    user = model.objects.get(pk=int(request.POST['pk']))
                    form = ManageProfileForm(request.POST)
                    if not form.is_valid():
                        raise FormError(form)
                    user.is_active=form.cleaned_data['user_is_active']
                    user.is_staff=form.cleaned_data['user_is_staff']
                    user.status = form.cleaned_data['user_is_active']
                    user.save(request)
                    administrativo = user.get_admin()
                    administrativo.status = form.cleaned_data['perfil_administrativo']
                    administrativo.save(request)
                    cliente = user.get_client(form.cleaned_data['perfil_cliente'])
                    cliente.status = form.cleaned_data['perfil_cliente']
                    cliente.save(request)
                    log(f"Modifico perfil usuario {user.username} - {user.get_full_name()}", request, "change")
                    messages.success(request, "Modificado correctamente.")
                    res_json.append({'error': False,
                                     "to": redirectAfterPostGet(request)
                                     })
                elif action == 'change_username':
                    user = model.objects.get(pk=int(request.POST['pk']))
                    form = ChangeUsernameForm(request.POST)
                    if not form.is_valid():
                        raise FormError(form)
                    if Usuario.objects.filter(username=form.cleaned_data['username']).exists():
                        raise ValueError("Ya existe un usuario con ese nombre de usuario.")
                    user.username = form.cleaned_data['username']
                    user.save(request)
                    log(f"Modifico nombre de usuario {user.username} - {user.get_full_name()}", request, "change")
                    messages.success(request, "Modificado correctamente.")
                    res_json.append({'error': False,
                                     "reload": True
                                     })
                elif action == 'uploadUsers':
                    form = CargarUsuariosForm(request.POST, request.FILES)
                    if form.is_valid():
                        archivo = request.FILES['archivo']
                        workbook = load_workbook(archivo)
                        sheet = workbook[workbook.sheetnames[0]]
                        linea_lectura = 1
                        creados = 0
                        editado = 0
                        for rowx in sheet.iter_rows(min_row=2):
                            cols = [cell.value for cell in rowx]

                            if len(cols) < 3:
                                continue

                            apellidos = (cols[0] or '').strip().upper()
                            nombres = (cols[1] or '').strip().upper()
                            cedula = str(cols[2] or '').strip()

                            if not (nombres and apellidos and cedula):
                                continue

                            telefono = str(cols[3] or '').strip()
                            email = (cols[4] or '').strip()

                            if not Usuario.objects.filter(username=cedula).exists():
                                user = Usuario.objects.create_user(
                                    username=cedula,
                                    email=email,
                                    password=cedula,
                                    first_name=nombres,
                                    last_name=apellidos,
                                    telefono=telefono,
                                    documento=cedula
                                )
                                creados += 1
                            else:
                                user = Usuario.objects.get(username=cedula)
                                user.documento = cedula
                                user.first_name = nombres
                                user.last_name = apellidos
                                user.telefono = telefono
                                user.email = email
                                editado += 1
                                user.status = True
                                user.is_active = True
                                user.save(request)
                            _success, resp = user.register_staff_user()
                            if not _success:
                                raise NameError(f"Error al registrar usuario {nombres} {apellidos}: {resp}")
                        log(f"Subió archivo de usuarios con {creados} creados y {editado} editados", request, "add")
                        messages.success(request, f"Creados: {creados}, Editados: {editado}")
                        res_json.append({'error': False, "reload": True})
                    else:
                        raise FormError(form)

        except ValueError as ex:
            res_json.append({'error': True,
                             "message": str(ex)
                             })
        except FormError as ex:
            res_json.append(ex.dict_error)
        except Exception as ex:
            res_json.append({'error': True,
                             "message": f"Intente Nuevamente {ex}"
                             })
        return JsonResponse(res_json, safe=False)
    elif request.method == 'GET':
        if 'action' in request.GET:
            data["action"] = action = request.GET['action']
            if action == 'add':
                form = Formulario()
                form.fields['provincia'].queryset = Provincia.objects.none()
                form.fields['ciudad'].queryset = Ciudad.objects.none()
                data["form"] = form
                return render(request, 'autenticacion/usuario/form.html', data)
            elif action == 'change':
                pk = int(request.GET['pk'])
                user = model.objects.get(pk=pk)
                data["pk"] = pk
                data["object"] = user
                form = Formulario(instance=user)
                if user.ciudad:
                    form.fields['provincia'].queryset = Provincia.objects.filter(status=True,
                                                                                 id=user.ciudad.provincia.id)
                    form.fields['ciudad'].queryset = Ciudad.objects.filter(status=True, id=user.ciudad.id)
                else:
                    form.fields['provincia'].queryset = Provincia.objects.none()
                    form.fields['ciudad'].queryset = Ciudad.objects.none()
                data["form"] = form
                return render(request, 'autenticacion/usuario/form.html', data)
            elif action == 'permisos_especiales':
                try:
                    data['id'] = id = int(request.GET['pk'])
                    data['filtro'] = filtro = Usuario.objects.get(pk=id)
                    data['permisos_especiales'] = _permisos_especiales_de(filtro)
                    titulo = f'Permisos especiales de {filtro.nombre_corto()}'
                    template = get_template("autenticacion/usuario/form_permisos_especiales.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'titulo': titulo})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})
            elif action == 'changegroup':
                try:
                    data['id'] = id = int(request.GET['pk'])
                    data['filtro'] = filtro = Usuario.objects.get(pk=id)
                    titulo= f'Gestionar Grupos de {filtro.nombre_corto()}'
                    form = GrupoUserForm(instance=filtro)
                    data['form'] = form
                    template = get_template("autenticacion/usuario/formmodal.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'titulo':titulo })
                except Exception as ex:
                    pass
            elif action == 'ver':
                pk = int(request.GET['pk'])
                user = model.objects.get(pk=pk)
                data["pk"] = pk
                data["object"] = user
                data["form"] = Formulario(instance=user, ver=True)
                return render(request, 'autenticacion/usuario/form.html', data)
            elif action =='manage_profile':
                data['id'] = id = int(request.GET['pk'])
                data['filtro'] = usuario = Usuario.objects.get(pk=id)
                titulo = f'Gestionar perfiles de {usuario.nombre_corto()}'
                form = ManageProfileForm()
                form.fields['perfil_administrativo'].initial = usuario.es_administrativo()
                form.fields['perfil_cliente'].initial = usuario.es_persona()
                form.fields['user_is_active'].initial = usuario.is_active
                form.fields['user_is_staff'].initial = usuario.is_staff
                data['form'] = form
                template = get_template("autenticacion/usuario/form_manage_profiles.html")
                return JsonResponse({"result": True, 'data': template.render(data), 'titulo': titulo})
            elif action =='change_username':
                try:
                    data['id'] = id = int(request.GET['pk'])
                    data['filtro'] = usuario = Usuario.objects.get(pk=id)
                    titulo = f'Cambiar nombre de usuario de {usuario.nombre_corto()}'
                    form = ChangeUsernameForm()
                    data['form'] = form
                    template = get_template("autenticacion/usuario/form_change_username.html")
                    return JsonResponse({"result": True, 'data': template.render(data), 'titulo': titulo})
                except Exception as ex:
                    return JsonResponse({"result": False, "message": f"Error: {ex}."})
            elif action == 'uploadUsers':
                try:
                    form = CargarUsuariosForm()
                    data['form'] = form
                    template = get_template("autenticacion/usuario/cargausuarios.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, "message": f"Error: {ex}."})
            elif action == 'exportar_excel':
                from django.http import HttpResponse
                from .funciones_usuario import filtros_listado_usuarios, exportar_usuarios_excel
                filtros_x = filtros_listado_usuarios(request.GET)
                listado_x = (
                    model.objects.filter(filtros_x)
                    .filter(perfiladministrativo__isnull=False)
                    .prefetch_related('groups')
                    .order_by('last_name')
                    .distinct()
                )
                wb = exportar_usuarios_excel(listado_x)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="usuarios_{date.today()}.xlsx"'
                wb.save(response)
                log(f"Exporto listado de usuarios a Excel ({listado_x.count()} filas)", request, "view")
                return response

        from .funciones_usuario import filtros_listado_usuarios

        grupoid, status_perfil, criterio, url_vars = request.GET.getlist('grupoid', ''), request.GET.get('status_perfil', ''), request.GET.get('criterio', ''), ''
        id, documento = request.GET.get('id', ''), request.GET.get('documento', '')
        orderby = request.GET.get('orderby', '')

        order = 'last_name'
        if orderby:
            data["orderby"] = orderby
            url_vars += '&orderby=' + orderby
            if orderby == '1':
                order = 'last_name'
            elif orderby == '2':
                order = '-last_name'
            elif orderby == '3':
                order = 'date_joined'
            elif orderby == '4':
                order = '-date_joined'

        if id:
            data['id'] = id
            url_vars += f'&id={id}'

        if documento:
            data['documento'] = documento
            url_vars += f'&documento={documento}'

        if status_perfil:
            data['status_perfil'] = status_perfil
            url_vars += f'&status_perfil={status_perfil}'

        if grupoid:
            data["grupoid"] = grupoid = list(map(lambda x: int(x), grupoid))
            for scl in grupoid:
                url_vars += "&grupoid={}".format(scl)

        if criterio:
            data['criterio'] = criterio
            url_vars += f'&criterio={criterio}'

        # El Q de filtros vive en funciones_usuario.filtros_listado_usuarios —
        # misma fuente que exportar_excel y change_password_masivo.
        filtros = filtros_listado_usuarios(request.GET)

        listado = model.objects.filter(filtros).filter(perfiladministrativo__isnull=False).order_by(order).distinct()
        data["url_vars"] = url_vars
        data["list_count"] = listado.count()
        data['gruposrol'] = Group.objects.all()
        paginador(request, listado, 20, data, url_vars)
        return render(request, 'autenticacion/usuario/listado.html', data)
