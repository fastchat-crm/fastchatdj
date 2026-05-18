import sys
import threading
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.forms import model_to_dict
from django.http import JsonResponse
from django.shortcuts import render
from django.template.loader import get_template
from openpyxl.reader.excel import load_workbook

from core.custom_models import FormError
from core.email_config import send_html_mail
from core.funciones import addData, paginador, secure_module, log, generar_nombre, renderizar_texto_dinamico
from .forms import CabMarketingMailingForm, MarketingMailSendForm, SendMailingForm
from .models import CabMarketingMailing, DetailMarketingMailing, TaskMarketingMail, DetailTaskMarketingMail
from django.contrib import messages
import time


@login_required
@secure_module
def mailingView(request):
    data = {
        'titulo': 'Email Administration',
        'modulo': 'Security',
        'ruta': request.path,
    }
    addData(request, data)
    persona = request.user
    if request.method == 'POST':
        res_json = []
        action = request.POST['action']
        try:
            with transaction.atomic():
                if action == 'addlista':
                    form = CabMarketingMailingForm(request.POST, request.FILES)
                    if form.is_valid():
                        cab_ = CabMarketingMailing(name=form.cleaned_data['name'])
                        cab_.save(request)
                        if 'file' not in request.FILES:
                            raise NameError('Error: no file uploaded')
                        newfile = request.FILES['file']
                        newfilesd = newfile._name
                        ext = newfilesd[newfilesd.rfind("."):]
                        if ext not in ('.xlsx', '.xls',):
                            raise NameError('Error: only .xls and .xlsx files allowed')
                        if newfile.size > 10485760:
                            raise NameError('Error: file exceeds 10 MB')
                        newfile._name = generar_nombre("lista_email", newfile._name)
                        cab_.file = newfile
                        cab_.save(request)
                        workbook = load_workbook(newfile)
                        sheet = workbook[workbook.sheetnames[0]]
                        total_lineas = sheet.max_row
                        linea_lectura = 1
                        for rowx in sheet.rows:
                            if linea_lectura > 1:
                                cols = [cell.value for cell in rowx]
                                document = cols[0]
                                last_name = cols[1]
                                first_name = cols[2]
                                email = cols[3]
                                emailinst = cols[4] if cols[4] else ''
                                email_copia = cols[5] if cols[5] else ''
                                if not document:
                                    raise NameError(f'Error: row {linea_lectura} is missing a document')
                                if not email:
                                    raise NameError(f'Error: row {linea_lectura} is missing an email')
                                detail = DetailMarketingMailing(cab=cab_, document=document, last_name=last_name, first_name=first_name, email=email, emailinst=emailinst, correo_copia=email_copia)
                                detail.save(request)
                            linea_lectura += 1
                        log(f"New mailing list registered: {cab_.__str__()}", request, "add")
                        messages.success(request, "Successfully registered")
                        res_json.append({'error': False, "reload": True})
                    else:
                        raise FormError(form)
                elif action == 'deletelista':
                    filtro = CabMarketingMailing.objects.get(pk=int(request.POST['id']))
                    filtro.status = False
                    filtro.save(request)
                    DetailMarketingMailing.objects.filter(cab=filtro).update(status=False)
                    TaskMarketingMail.objects.filter(cab=filtro).update(status=False)
                    DetailTaskMarketingMail.objects.filter(task__cab=filtro).update(status=False)
                    log(f"Mailing list deleted: {filtro.__str__()}", request, "delete")
                    messages.success(request, "Record deleted")
                    res_json = {"error": False}
                elif action == 'addtarea':
                    form = MarketingMailSendForm(request.POST, request.FILES)
                    if form.is_valid():
                        task = TaskMarketingMail(
                            cab=form.cleaned_data['cab'],
                            title=form.cleaned_data['title'],
                            body=form.cleaned_data['body'],
                            envia_copia=form.cleaned_data['envia_copia'],
                            correo_copia=form.cleaned_data['correo_copia'],
                        )
                        task.save(request)
                        if 'image' in request.FILES:
                            newfile = request.FILES['image']
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext not in (".jpg", ".png", ".jpeg"):
                                raise NameError('Error: only jpg, png, jpeg files allowed')
                            if newfile.size > 16194304:
                                raise NameError('Error: file exceeds 4 MB')
                            newfile._name = generar_nombre("lista_email", newfile._name)
                            task.image = newfile
                            task.upload_image = True
                            task.save(request)
                        log(f"New send task registered: {task.__str__()}", request, "add")
                        messages.success(request, "Successfully registered")
                        res_json.append({'error': False, "reload": True})
                    else:
                        raise FormError(form)
                elif action == 'changetarea':
                    filtro = TaskMarketingMail.objects.get(pk=int(request.POST['pk']))
                    form = MarketingMailSendForm(request.POST, request.FILES)
                    if form.is_valid():
                        filtro.cab = form.cleaned_data['cab']
                        filtro.title = form.cleaned_data['title']
                        filtro.body = form.cleaned_data['body']
                        filtro.envia_copia = form.cleaned_data['envia_copia']
                        filtro.correo_copia = form.cleaned_data['correo_copia']
                        if 'image' in request.FILES:
                            newfile = request.FILES['image']
                            newfilesd = newfile._name
                            ext = newfilesd[newfilesd.rfind("."):]
                            if ext not in (".jpg", ".png", ".jpeg"):
                                raise NameError('Error: only jpg, png, jpeg files allowed')
                            if newfile.size > 16194304:
                                raise NameError('Error: file exceeds 4 MB')
                            newfile._name = generar_nombre("lista_email", newfile._name)
                            filtro.image = newfile
                        filtro.save(request)
                        log(f"Send task updated: {filtro.__str__()}", request, "add")
                        messages.success(request, "Successfully updated")
                        res_json.append({'error': False, "reload": True})
                    else:
                        raise FormError(form)
                elif action == 'deletetarea':
                    filtro = TaskMarketingMail.objects.get(pk=int(request.POST['id']))
                    filtro.status = False
                    filtro.save(request)
                    DetailTaskMarketingMail.objects.filter(task=filtro).update(status=False)
                    log(f"Send task deleted: {filtro.__str__()}", request, "delete")
                    messages.success(request, "Record deleted")
                    res_json = {"error": False}
                elif action == 'sendlistmailing':
                    filtro = TaskMarketingMail.objects.get(pk=int(request.POST['pk']))
                    form = SendMailingForm(request.POST)
                    if not form.is_valid():
                        raise FormError(form)
                    if filtro:
                        SendMarketingEmail(request, filtro).start()
                        log(f"Mass email send started: {filtro.__str__()}", request, "add")
                        filtro.sent_status = True
                        filtro.sent_date = datetime.now()
                        filtro.sent_user = persona
                        filtro.save(request)
                        messages.success(request, f"Sending emails. You will be notified at {request.user.email} when the process finishes.")
                        res_json.append({'error': False, "reload": True})
                    else:
                        raise NameError('Error: task not found')
        except ValueError as ex:
            res_json.append({'error': True, "message": str(ex)})
        except FormError as ex:
            res_json.append(ex.dict_error)
        except Exception as ex:
            eline = 'Error on line {} - {}'.format(sys.exc_info()[-1].tb_lineno, ex)
            res_json.append({'error': True, "message": f"Please try again: {ex}, {eline}"})
        return JsonResponse(res_json, safe=False)
    elif request.method == 'GET':
        if 'action' in request.GET:
            data["action"] = action = request.GET['action']
            if action == 'addlista':
                try:
                    form = CabMarketingMailingForm()
                    data['form'] = form
                    template = get_template("seguridad/mailing/listas/form.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass
            elif action == 'detalle':
                try:
                    data['filtro'] = filtro = CabMarketingMailing.objects.get(pk=int(request.GET['id']))
                    data['subtitulo'] = f'List Detail: {filtro.__str__()}'
                    url_vars = f'&action={action}&id={filtro.id}'
                    listado = filtro.list_details()
                    data["list_count"] = listado.count()
                    data["url_vars"] = url_vars
                    paginador(request, listado, 20, data, url_vars)
                    return render(request, 'seguridad/mailing/listas/view_lista_detalle.html', data)
                except Exception as ex:
                    pass
            elif action == 'tareas':
                try:
                    data['subtitulo'] = 'Send Tasks'
                    criterio, filtros, url_vars = request.GET.get('criterio', '').strip(), Q(status=True), f'&action={action}'
                    id = request.GET.get('id', '')
                    if criterio:
                        filtros = filtros & (Q(title__icontains=criterio))
                        data["criterio"] = criterio
                        url_vars += '&criterio=' + criterio
                    if id:
                        filtros = filtros & (Q(pk=int(id)))
                        data["id"] = id
                        url_vars += '&id=' + id
                    listado = TaskMarketingMail.objects.filter(filtros)
                    if not request.user.is_superuser:
                        listado = listado.filter(usuario_creacion=request.user)
                    data["list_count"] = listado.count()
                    data["url_vars"] = url_vars
                    request.session['viewthpage'] = 2
                    paginador(request, listado.order_by('-id'), 20, data, url_vars)
                    return render(request, 'seguridad/mailing/envios/listado.html', data)
                except Exception as ex:
                    pass
            elif action == 'addtarea':
                try:
                    form = MarketingMailSendForm()
                    if not request.user.is_superuser:
                        form.fields['cab'].queryset = CabMarketingMailing.objects.filter(status=True, usuario_creacion=request.user)
                    data['form'] = form
                    template = get_template("seguridad/mailing/envios/form.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})
            elif action == 'changetarea':
                try:
                    filtro = TaskMarketingMail.objects.get(pk=int(request.GET['id']))
                    form = MarketingMailSendForm(initial=model_to_dict(filtro))
                    form.fields['cab'].queryset = CabMarketingMailing.objects.filter(pk=filtro.cab.pk)
                    data['form'] = form
                    data['filtro'] = filtro
                    if filtro.image:
                        form.fields['image'].widget.attrs['data-default-file'] = filtro.image.url
                    template = get_template("seguridad/mailing/envios/form.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})
            elif action == 'detalletareas':
                try:
                    data['filtro'] = filtro = TaskMarketingMail.objects.get(pk=int(request.GET['id']))
                    data['subtitulo'] = f'Send Detail: {filtro.__str__()}'
                    url_vars = f'&action={action}&id={filtro.id}'
                    listado = filtro.list_details()
                    data["list_count"] = listado.count()
                    data["url_vars"] = url_vars
                    paginador(request, listado, 20, data, url_vars)
                    return render(request, 'seguridad/mailing/envios/view_envios_detalle.html', data)
                except Exception as ex:
                    pass
            elif action == 'sendlistmailing':
                try:
                    data['filtro'] = filtro = TaskMarketingMail.objects.get(pk=int(request.GET['id']))
                    data['form'] = SendMailingForm()
                    template = get_template("seguridad/mailing/envios/form_send.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'message': str(ex)})

        data['subtitulo'] = 'Mailing Lists'
        criterio, filtros, url_vars = request.GET.get('criterio', '').strip(), Q(status=True), f''
        if criterio:
            filtros = filtros & (Q(name__icontains=criterio))
            data["criterio"] = criterio
            url_vars += '&criterio=' + criterio
        listado = CabMarketingMailing.objects.filter(filtros)
        if not request.user.is_superuser:
            listado = listado.filter(usuario_creacion=request.user)
        data["list_count"] = listado.count()
        data["url_vars"] = url_vars
        request.session['viewthpage'] = 1
        paginador(request, listado.order_by('-id'), 20, data, url_vars)
        return render(request, 'seguridad/mailing/listas/listado.html', data)


class SendMarketingEmail(threading.Thread):
    def __init__(self, request, filtro):
        self.request = request
        self.filtro = filtro
        threading.Thread.__init__(self)

    def run(self):
        request, filtro = self.request, self.filtro
        persona = request.user
        try:
            filtro_ = TaskMarketingMail.objects.get(pk=filtro.pk)
            if not filtro_:
                raise NameError('Error: task not found')
            list_send = DetailMarketingMailing.objects.filter(cab=filtro_.cab, status=True)
            count = 0
            lista_cc = [filtro_.correo_copia] if filtro_.envia_copia else []
            for mail in list_send:
                list_email = mail.email_list()
                if mail.correo_copia:
                    lista_cc.append(mail.correo_copia)
                if not DetailTaskMarketingMail.objects.filter(task=filtro_, email=mail.email, status=True).exists():
                    DetailTaskMarketingMail.objects.create(task=filtro_, detail=mail, document=mail.document, last_name=mail.last_name, first_name=mail.first_name, email=mail.email, emailinst=mail.emailinst, notified=True)
                texto_renderizado = renderizar_texto_dinamico(filtro_.body, {'nombres': mail.datos()})
                datos_ = {'title': filtro_.title, 'mensaje': texto_renderizado}
                adjunto_img = [filtro_.image] if filtro_.upload_image else []
                send_html_mail(filtro_.title, "email/email_default.html", datos_, list_email, lista_cc, adjuntossave=adjunto_img)
                mail.notified = True
                mail.date_notified = datetime.now()
                mail.user_notified = persona
                mail.save(request)
                if count % 30 == 0:
                    time.sleep(30)
                else:
                    time.sleep(2)
                count += 1
            datos_ = {'title': 'MAILING SEND SUCCESS', 'mensaje': f"Successfully sent {count} emails for mailing list '{filtro_.cab.name}'."}
            send_html_mail('MAILING SEND SUCCESS', "email/email_default.html", datos_, [persona.email] if persona.email else [], [])
            return True, ""
        except Exception as ex:
            linea_ = 'Error on line {}'.format(sys.exc_info()[-1].tb_lineno)
            datos_ = {'title': 'MAILING SEND ERROR', 'mensaje': f"{ex}, {linea_}"}
            send_html_mail('MAILING SEND ERROR', "email/email_default.html", datos_, [persona.email] if persona.email else [], [])
            return False, f"{ex} - {linea_}"
