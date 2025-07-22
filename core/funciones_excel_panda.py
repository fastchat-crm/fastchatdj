import pandas as pd
from django.db import connection
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def export_query_to_excel(query, params, file_name, sheet_name='Reporte'):
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            data = cursor.fetchall()

        df = pd.DataFrame(data, columns=columns)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Define styles
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_font = Font(bold=True)
            alignment = Alignment(horizontal='center', vertical='center')

            # Apply styles to header
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
                cell.border = border_style
                # Adjust column width
                worksheet.column_dimensions[get_column_letter(col_num)].width = 20

            # Apply styles to data rows
            for row_num, row in enumerate(
                    worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)),
                    start=2):
                for cell in row:
                    # Apply border to each cell
                    cell.border = border_style

                    # Apply center alignment to all cells
                    cell.alignment = alignment

                    # Example: Apply different fills based on value (e.g., highlight negative values)
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                fill_type="solid")  # Red fill for negative numbers
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                                                fill_type="solid")  # White fill for other cells

        return response
    except Exception as e:
        res_json = {'error': True, 'message': f'{e}'}
        return JsonResponse(res_json, safe=False)

